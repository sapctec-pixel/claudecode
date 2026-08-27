# -*- coding: utf-8 -*-
"""بناء نموذج حساب معايير القوى العاملة — الجزء 1: الأوراق المعيارية ومحرك الحساب."""
import sys
sys.path.insert(0, '/home/user/claudecode/moh-workforce')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.comments import Comment

import data_general_doctors as gd
import data_general_rest as gr
import data_maternity as mt
import data_mh_ltc as ml

FONT = "Arial"
NAVY   = "1F3864"
TEAL   = "1F6F6B"
GOLD   = "BF8F00"
LIGHT  = "EEF3F8"
INPUTF = "FFF2CC"
GREY   = "F2F2F2"
WHITE  = "FFFFFF"

thin = Side(style="thin", color="BFBFBF")
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)

def h1(ws, text, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=text)
    c.font = Font(name=FONT, size=15, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

def h2(ws, text, ncols):
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    c = ws.cell(row=2, column=2 if False else 1, value=text)
    c.font = Font(name=FONT, size=9, italic=True, color="595959")
    c.fill = PatternFill("solid", fgColor=LIGHT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

HDR_FILL = PatternFill("solid", fgColor=TEAL)
HDR_FONT = Font(name=FONT, size=10, bold=True, color=WHITE)
TIER_FILL = PatternFill("solid", fgColor="D9E2F3")
CALC_FILL = PatternFill("solid", fgColor="E2F0D9")
IN_FILL   = PatternFill("solid", fgColor=INPUTF)
IN_FONT   = Font(name=FONT, size=10, color="0000FF")

CATS = ["طبيب", "تمريض", "قبالة", "أخصائي غير طبيب", "فني", "صيدلي", "مساعد صحي", "إداري"]


def build_standard_sheet(wb, title, subtitle, rows, tiers, published=None, note_extra=None):
    """ينشئ ورقة معيار: أعمدة الشرائح + الاحتياج المحسوب + الموجود + الفجوة."""
    ws = wb.create_sheet(title)
    ws.sheet_view.rightToLeft = True
    n = len(tiers)
    c_t0 = 4                       # أول عمود شريحة (D)
    c_t1 = 3 + n                   # آخر عمود شريحة
    L_t0, L_t1 = get_column_letter(c_t0), get_column_letter(c_t1)
    c_need = c_t1 + 1              # الاحتياج المحسوب
    c_act  = c_t1 + 2              # الموجود فعلياً
    c_gap  = c_t1 + 3              # الفجوة
    c_cov  = c_t1 + 4              # نسبة التغطية
    c_st   = c_t1 + 5              # الحالة
    c_raw  = c_t1 + 6              # قبل التقريب
    c_note = c_t1 + 7              # ملاحظات
    ncols = c_note

    h1(ws, title, ncols)
    h2(ws, subtitle, ncols)

    hdr = 4
    heads = ["م", "الفئة", "المسمى الوظيفي"] + [f"{t} سرير" for t in tiers] + \
            ["الاحتياج للسعة المدخلة", "الموجود فعلياً", "الفجوة (+فائض / -عجز)",
             "نسبة التغطية", "الحالة", "القيمة قبل التقريب", "ملاحظات"]
    for j, t in enumerate(heads, start=1):
        c = ws.cell(row=hdr, column=j, value=t)
        c.font = HDR_FONT; c.fill = HDR_FILL; c.border = BOX
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[hdr].height = 42

    # معاملات الاستيفاء (تُحسب مرة واحدة لكل ورقة بدل تكرارها في كل صف)
    ws.cell(row=3, column=3, value="معاملا الاستيفاء ◄").font = Font(
        name=FONT, size=8, color="A6A6A6")
    ws.cell(row=3, column=3).alignment = Alignment(horizontal="left")
    ci = ws.cell(row=3, column=4,
                 value=f"=IFERROR(MATCH(CAP,${L_t0}${hdr}:${L_t1}${hdr},1),1)")
    ci.font = Font(name=FONT, size=8, color="A6A6A6"); ci.number_format = "0"
    cwt = ws.cell(row=3, column=5, value=(
        f"=IF(OR(CAP<=${L_t0}${hdr},CAP>=${L_t1}${hdr}),0,"
        f"(CAP-INDEX(${L_t0}${hdr}:${L_t1}${hdr},$D$3))"
        f"/(INDEX(${L_t0}${hdr}:${L_t1}${hdr},$D$3+1)-INDEX(${L_t0}${hdr}:${L_t1}${hdr},$D$3)))"))
    cwt.font = Font(name=FONT, size=8, color="A6A6A6"); cwt.number_format = "0.000"

    r0 = hdr + 1
    for i, (cat, name, vals, pub_total, note) in enumerate(rows):
        r = r0 + i
        ws.cell(row=r, column=1, value=i + 1).font = Font(name=FONT, size=9)
        ws.cell(row=r, column=2, value=cat).font = Font(name=FONT, size=9, color="595959")
        c = ws.cell(row=r, column=3, value=name)
        c.font = Font(name=FONT, size=10)
        c.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
        for j, v in enumerate(vals):
            cc = ws.cell(row=r, column=c_t0 + j, value=v)
            cc.font = Font(name=FONT, size=10)
            cc.fill = TIER_FILL
            cc.number_format = "#,##0"
            cc.alignment = Alignment(horizontal="center")
        # الاستيفاء الخطي بين الشرائح (يعتمد على معاملَي الاستيفاء المحسوبين في الصف 3)
        raw = (
            f"=IF(CAP<=${L_t0}${hdr},{L_t0}{r}*CAP/${L_t0}${hdr},"
            f"IF(CAP>=${L_t1}${hdr},{L_t1}{r}*CAP/${L_t1}${hdr},"
            f"INDEX(${L_t0}{r}:${L_t1}{r},$D$3)"
            f"+$E$3*(INDEX(${L_t0}{r}:${L_t1}{r},$D$3+1)-INDEX(${L_t0}{r}:${L_t1}{r},$D$3))))"
        )
        L_raw = get_column_letter(c_raw)
        L_need = get_column_letter(c_need)
        L_act = get_column_letter(c_act)
        cr = ws.cell(row=r, column=c_raw, value=raw)
        cr.font = Font(name=FONT, size=9, color="A6A6A6"); cr.number_format = "0.00"
        cn = ws.cell(row=r, column=c_need,
                     value=f'=IF(RMODE="تقريب لأعلى",ROUNDUP({L_raw}{r},0),ROUND({L_raw}{r},0))')
        cn.font = Font(name=FONT, size=10, bold=True); cn.fill = CALC_FILL
        cn.number_format = "#,##0"; cn.border = BOX
        cn.alignment = Alignment(horizontal="center")
        ca = ws.cell(row=r, column=c_act)
        ca.fill = IN_FILL; ca.font = IN_FONT; ca.number_format = "#,##0"; ca.border = BOX
        ca.alignment = Alignment(horizontal="center")
        cg = ws.cell(row=r, column=c_gap, value=f"={L_act}{r}-{L_need}{r}")
        cg.number_format = "#,##0;[Red]-#,##0;-"; cg.font = Font(name=FONT, size=10)
        cg.alignment = Alignment(horizontal="center")
        cv = ws.cell(row=r, column=c_cov, value=f'=IF({L_need}{r}=0,"",{L_act}{r}/{L_need}{r})')
        cv.number_format = "0%"; cv.font = Font(name=FONT, size=10)
        cv.alignment = Alignment(horizontal="center")
        cs = ws.cell(row=r, column=c_st,
                     value=f'=IF({L_need}{r}=0,"لا ينطبق",IF({L_act}{r}>={L_need}{r},"مكتمل",'
                           f'IF({L_act}{r}=0,"لا يوجد","عجز")))')
        cs.font = Font(name=FONT, size=10); cs.alignment = Alignment(horizontal="center")
        cno = ws.cell(row=r, column=c_note, value=note)
        cno.font = Font(name=FONT, size=8, color="7F7F7F")
        cno.alignment = Alignment(horizontal="right", wrap_text=True)

    rlast = r0 + len(rows) - 1
    rt = rlast + 1
    ws.cell(row=rt, column=3, value="المجموع (محسوب من الجدول)")
    for j in range(c_t0, c_note):
        L = get_column_letter(j)
        if j in (c_cov, c_st):
            continue
        ws.cell(row=rt, column=j, value=f"=SUM({L}{r0}:{L}{rlast})")
    Lneed, Lact = get_column_letter(c_need), get_column_letter(c_act)
    ws.cell(row=rt, column=c_cov, value=f'=IF({Lneed}{rt}=0,"",{Lact}{rt}/{Lneed}{rt})')
    for j in range(1, ncols + 1):
        c = ws.cell(row=rt, column=j)
        c.font = Font(name=FONT, size=10, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=NAVY); c.border = BOX
        if j >= c_t0: c.alignment = Alignment(horizontal="center")
        if j in range(c_t0, c_need + 1) or j == c_gap: c.number_format = "#,##0"
    ws.cell(row=rt, column=c_cov).number_format = "0%"

    # صف المقارنة مع المجموع المنشور في الدليل
    if published:
        rp = rt + 1
        ws.cell(row=rp, column=3, value="المجموع المنشور في الدليل")
        for j, v in enumerate(published):
            ws.cell(row=rp, column=c_t0 + j, value=v)
        rd = rp + 1
        ws.cell(row=rd, column=3, value="الفرق (محسوب − منشور)")
        for j in range(len(published)):
            L = get_column_letter(c_t0 + j)
            ws.cell(row=rd, column=c_t0 + j, value=f"={L}{rt}-{L}{rp}")
        for rr, fill in ((rp, GREY), (rd, GREY)):
            for j in range(1, c_t1 + 1):
                c = ws.cell(row=rr, column=j)
                c.font = Font(name=FONT, size=9, italic=True)
                c.fill = PatternFill("solid", fgColor=fill)
                c.alignment = Alignment(horizontal="center")
                c.number_format = "#,##0"
        if note_extra:
            ws.cell(row=rd, column=c_t1 + 1, value=note_extra).font = Font(
                name=FONT, size=8, color="C00000")

    # تنسيق شرطي على عمود الحالة
    Lst = get_column_letter(c_st)
    rng = f"{Lst}{r0}:{Lst}{rlast}"
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=[f'OR({Lst}{r0}="عجز",{Lst}{r0}="لا يوجد")'],
        fill=PatternFill("solid", fgColor="F8CBAD"), font=Font(name=FONT, color="833C0B")))
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=[f'{Lst}{r0}="مكتمل"'],
        fill=PatternFill("solid", fgColor="C6E0B4"), font=Font(name=FONT, color="375623")))

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 46
    for j in range(c_t0, c_t1 + 1):
        ws.column_dimensions[get_column_letter(j)].width = 9
    ws.column_dimensions[get_column_letter(c_need)].width = 13
    ws.column_dimensions[get_column_letter(c_act)].width = 12
    ws.column_dimensions[get_column_letter(c_gap)].width = 13
    ws.column_dimensions[get_column_letter(c_cov)].width = 10
    ws.column_dimensions[get_column_letter(c_st)].width = 11
    ws.column_dimensions[get_column_letter(c_raw)].width = 11
    ws.column_dimensions[get_column_letter(c_note)].width = 42
    ws.freeze_panes = ws.cell(row=r0, column=c_t0)
    ws.sheet_properties.tabColor = TEAL
    return dict(sheet=title, r0=r0, rlast=rlast, rtot=rt,
                col_cat="B", col_need=get_column_letter(c_need),
                col_act=get_column_letter(c_act), tiers=tiers)
