# -*- coding: utf-8 -*-
import sys
sys.path.insert(0, '/home/user/claudecode/moh-workforce')
from build_part1 import *
from openpyxl.workbook.defined_name import DefinedName

OUT = "/home/user/claudecode/moh-workforce/MOH_Workforce_Standards_2023.xlsx"
SRC = "المصدر: المعايير القياسية للقوى العاملة بمستشفيات وزارة الصحة — النسخة الثانية 2023م — وكالة الوزارة للخدمات العلاجية / الإدارة العامة لشؤون المستشفيات"

TYPES = ["مستشفى عام", "مستشفى ولادة وأطفال", "مجمع إرادة والصحة النفسية", "مستشفى رعاية مديدة"]

wb = Workbook()
wb.remove(wb.active)

# ======================= 1) دليل الاستخدام =======================
ws = wb.create_sheet("دليل الاستخدام"); ws.sheet_view.rightToLeft = True
h1(ws, "دليل استخدام نموذج حساب معايير القوى العاملة", 4)
h2(ws, SRC, 4)
guide = [
 ("", ""),
 ("كيف يعمل النموذج؟", ""),
 ("1", "افتح ورقة «لوحة التحكم» وأدخل: اسم المستشفى، نوع المستشفى، والسعة السريرية (عدد الأسرة)."),
 ("2", "تُحتسب أعمدة «الاحتياج للسعة المدخلة» في كل ورقة معيار تلقائياً بمجرد إدخال السعة السريرية."),
 ("3", "أدخل الأعداد الحالية للموظفين في عمود «الموجود فعلياً» (الخلايا الصفراء) في أوراق المعايير."),
 ("4", "تظهر الفجوة ونسبة التغطية والحالة تلقائياً لكل مسمى وظيفي، ويتجمّع الملخص في «لوحة التحكم»."),
 ("", ""),
 ("قواعد الحساب", ""),
 ("الشرائح المعيارية", "الدليل يحدد الأعداد عند سعات محددة (50 / 100 / 150 / 200 / 300 / 400 / 500 سرير للمستشفيات العامة)."),
 ("السعات غير المذكورة", "يُستخدم الاستيفاء الخطي (Linear Interpolation) بين أقرب شريحتين. مثال: سعة 250 سرير = القيمة عند 200 + نصف الفرق بين 200 و 300."),
 ("أقل من أصغر شريحة", "يُحسب تناسبياً من أصغر شريحة (القيمة × السعة ÷ أصغر شريحة)."),
 ("أكبر من أكبر شريحة", "يُحسب تناسبياً من أكبر شريحة (القيمة × السعة ÷ أكبر شريحة)."),
 ("التقريب", "افتراضياً «تقريب لأعلى» لأن الدليل يمثل الحد الأدنى للقوى العاملة. يمكن تغييره من لوحة التحكم إلى «تقريب لأقرب صحيح»."),
 ("عند السعات المعيارية", "تُطابق نتائج النموذج أرقام الدليل حرفياً (تم التحقق من كل جدول مقابل المجاميع المنشورة)."),
 ("", ""),
 ("دلالة الألوان", ""),
 ("أصفر / خط أزرق", "خلايا إدخال — أنت من يعبّئها."),
 ("أخضر فاتح", "خلايا محسوبة تلقائياً — لا تُعدَّل."),
 ("أزرق فاتح", "قيم المعيار كما وردت في الدليل — مرجعية ثابتة."),
 ("رمادي", "صفوف مقارنة/تدقيق."),
 ("", ""),
 ("تنبيهات مهنية", ""),
 ("الحد الأدنى", "المعيار يمثل الحد الأدنى للقوى العاملة وليس العدد الأمثل (إيضاح 1)."),
 ("نطاق الخدمة", "لا يُطبَّق معيار أي تخصص ما لم تتوفر خدمته فعلياً في المنشأة (إيضاح 42) — راجع ورقة «نطاق الخدمات»."),
 ("العنايات المركزة", "يُطبَّق معيار العنايات المركزة فقط إذا كانت المستشفى تقدم الخدمة ومعتمدة (إيضاح 8)."),
 ("الأقسام المشروطة", "الأورام، زراعة القوقعة، العقم، جراحة السمنة، التأهيل الطبي — تُطبَّق عند اعتماد المركز فقط (إيضاحات 34، 37-41)."),
 ("راجع الإيضاحات", "ورقة «الإيضاحات» تحتوي الإيضاحات الـ 42 الواردة في الدليل كاملة."),
]
r = 4
for a, b in guide:
    if b == "" and a != "":
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        c = ws.cell(row=r, column=1, value=a)
        c.font = Font(name=FONT, size=12, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=TEAL)
        c.alignment = Alignment(horizontal="center")
        ws.row_dimensions[r].height = 24
    elif a == "" and b == "":
        pass
    else:
        ws.cell(row=r, column=1, value=a).font = Font(name=FONT, size=10, bold=True, color=NAVY)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        c = ws.cell(row=r, column=2, value=b)
        c.font = Font(name=FONT, size=10)
        c.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
        ws.row_dimensions[r].height = 30
    r += 1
ws.column_dimensions["A"].width = 24
for L in "BCD": ws.column_dimensions[L].width = 34
ws.sheet_properties.tabColor = GOLD

# ======================= 2) لوحة التحكم =======================
dash = wb.create_sheet("لوحة التحكم"); dash.sheet_view.rightToLeft = True
h1(dash, "لوحة التحكم — حساب الاحتياج من القوى العاملة", 8)
h2(dash, SRC, 8)

dash.merge_cells("B4:H4")
c = dash.cell(row=4, column=2, value="① بيانات المستشفى — عبّئ الخلايا الصفراء")
c.font = Font(name=FONT, size=12, bold=True, color=WHITE); c.fill = PatternFill("solid", fgColor=TEAL)
c.alignment = Alignment(horizontal="center"); dash.row_dimensions[4].height = 24

inputs = [
 (5, "اسم المستشفى", "مستشفى الملك فهد العام (مثال — استبدله)"),
 (6, "المنطقة / التجمع الصحي", "تجمع ... الصحي (مثال — استبدله)"),
 (7, "نوع المستشفى", "مستشفى عام"),
 (8, "السعة السريرية (عدد الأسرة)", 250),
 (9, "طريقة التقريب", "تقريب لأعلى"),
 (10, "اعتماد الأسرة المتحركة", "لا"),
 (11, "تاريخ التقييم", "1447/01/01هـ"),
]
for row, lab, val in inputs:
    cl = dash.cell(row=row, column=2, value=lab)
    cl.font = Font(name=FONT, size=11, bold=True); cl.fill = PatternFill("solid", fgColor=LIGHT)
    cl.border = BOX; cl.alignment = Alignment(horizontal="right", vertical="center")
    dash.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
    cv = dash.cell(row=row, column=3, value=val)
    cv.fill = IN_FILL; cv.font = Font(name=FONT, size=11, bold=True, color="0000FF")
    cv.border = BOX; cv.alignment = Alignment(horizontal="center", vertical="center")
    dash.row_dimensions[row].height = 22
dash.cell(row=8, column=3).number_format = "#,##0"

dv_type = DataValidation(type="list", formula1='"' + ",".join(TYPES) + '"', allow_blank=False)
dash.add_data_validation(dv_type); dv_type.add(dash["C7"])
dv_round = DataValidation(type="list", formula1='"تقريب لأعلى,تقريب لأقرب صحيح"', allow_blank=False)
dash.add_data_validation(dv_round); dv_round.add(dash["C9"])
dv_yn = DataValidation(type="list", formula1='"نعم,لا"', allow_blank=False)
dash.add_data_validation(dv_yn); dv_yn.add(dash["C10"])

dash["E7"] = "◄ اختر من القائمة"; dash["E8"] = "◄ أي رقم؛ يُستوفى خطياً بين شرائح الدليل"
dash["E9"] = "◄ الدليل حد أدنى ⇒ الافتراضي: تقريب لأعلى"
dash["E10"] = "◄ إيضاحات 26-29: تؤثر على نسبة التمريض لكل سرير"
for rr in (7, 8, 9, 10):
    dash.cell(row=rr, column=5).font = Font(name=FONT, size=9, italic=True, color="7F7F7F")

wb.defined_names.add(DefinedName("CAP", attr_text="'لوحة التحكم'!$C$8"))
wb.defined_names.add(DefinedName("HTYPE", attr_text="'لوحة التحكم'!$C$7"))
wb.defined_names.add(DefinedName("RMODE", attr_text="'لوحة التحكم'!$C$9"))
wb.defined_names.add(DefinedName("MOBILEBEDS", attr_text="'لوحة التحكم'!$C$10"))

# ======================= 3) أوراق المعايير =======================
T7 = gd.TIERS
T6M = mt.TIERS_M
T6MH = ml.TIERS_MH
T4 = ml.TIERS_LTC
meta = {}

meta["gd"] = build_standard_sheet(wb, "عام - الأطباء",
    "الملحق (1) — تفاصيل المعيار المحدث لفئة الأطباء (صفحات 22-25)", gd.DOCTORS, T7,
    gd.PUBLISHED_DOCTORS_TOTALS)
meta["gs"] = build_standard_sheet(wb, "عام - أخصائي غير طبيب",
    "الملحق (1) — أخصائي غير طبيب، ويشمل التمريض والقبالة (صفحات 26-27)", gr.SPECIALISTS, T7,
    gr.PUB_SPEC_TOTAL)
meta["gt"] = build_standard_sheet(wb, "عام - الفنيون",
    "الملحق (1) — الفنيون، ويشمل فني تمريض وفني قبالة (صفحة 28)", gr.TECHNICIANS, T7,
    gr.PUB_TECH_TOTAL)
meta["gp"] = build_standard_sheet(wb, "عام - الصيادلة",
    "الملحق (1) — الصيادلة (صفحة 29)", gr.PHARMACISTS, T7, gr.PUB_PHARM_TOTAL)
meta["ga"] = build_standard_sheet(wb, "عام - المساعد الصحي",
    "الملحق (1) — المساعد الصحي (صفحة 29)", gr.AUX, T7, gr.PUB_AUX_TOTAL)
meta["gadm"] = build_standard_sheet(wb, "عام - الإداريون",
    "الملحق (1) — الإداريون (صفحات 29-30)", gr.ADMIN, T7, gr.PUB_ADMIN_TOTAL)

meta["md"] = build_standard_sheet(wb, "ولادة - الأطباء",
    "الملحق (2) — مستشفيات الولادة والأطفال: فئة الأطباء (صفحات 48-50)", mt.DOCTORS_M, T6M,
    mt.PUB_DOCTORS_M)
meta["ms"] = build_standard_sheet(wb, "ولادة - أخصائي غير طبيب",
    "الملحق (2) — أخصائي غير طبيب، ويشمل التمريض والقبالة (صفحات 50-52)", mt.SPECIALISTS_M, T6M)
meta["mt"] = build_standard_sheet(wb, "ولادة - الفنيون",
    "الملحق (2) — الفنيون، ويشمل فني تمريض وفني قبالة (صفحات 52-53)", mt.TECHNICIANS_M, T6M,
    mt.PUB_TECH_M_TOTAL,
    note_extra="الفرق = صف «فني ترميز طبي» الوارد في جدول الدليل دون رقم تسلسلي وغير مُدرج في مجموعه المنشور.")
meta["mp"] = build_standard_sheet(wb, "ولادة - الصيادلة",
    "الملحق (2) — الصيادلة (صفحة 53)", mt.PHARMACISTS_M, T6M, mt.PUB_PHARM_M)
meta["madm"] = build_standard_sheet(wb, "ولادة - الإداريون",
    "الملحق (2) — الإداريون (صفحات 53-54)", mt.ADMIN_M, T6M, mt.PUB_ADMIN_M)

meta["mh"] = build_standard_sheet(wb, "الصحة النفسية وإرادة",
    "الملحق (3) — تفاصيل المعيار المحدّث لمجمعات إرادة والصحة النفسية (صفحات 58-59)",
    ml.MH, T6MH, ml.PUB_MH_TOTAL)
meta["lt"] = build_standard_sheet(wb, "الرعاية المديدة - الفنية",
    "الملحق (4) — مستشفيات الرعاية المديدة: الأقسام الفنية (صفحات 61-63)",
    ml.LTC_TECH, T4, ml.PUB_LTC_TECH)
meta["la"] = build_standard_sheet(wb, "الرعاية المديدة - الإداريون",
    "الملحق (4) — مستشفيات الرعاية المديدة: الوظائف الإدارية (صفحات 63-64)",
    ml.LTC_ADMIN, T4, ml.PUB_LTC_ADMIN)

# صف المعيار النسبي في الرعاية المديدة (أخصائي علاج تنفسي: أخصائي لكل 8 مرضى)
wl = wb["الرعاية المديدة - الفنية"]
rx = meta["lt"]["rtot"] + 4
wl.cell(row=rx, column=3, value="معيار نسبي إضافي (غير مُدرج في المجموع أعلاه):").font = Font(
    name=FONT, size=10, bold=True, color=NAVY)
wl.cell(row=rx + 1, column=3, value="أخصائي علاج تنفسي — أخصائي لكل 8 مرضى").font = Font(name=FONT, size=10)
cc = wl.cell(row=rx + 1, column=meta["lt"]["tiers"].__len__() + 4, value="=ROUNDUP(CAP/8,0)")
cc.font = Font(name=FONT, size=10, bold=True); cc.fill = CALC_FILL; cc.number_format = "#,##0"
cc.border = BOX; cc.alignment = Alignment(horizontal="center")


# ======================= 4) مصفوفة الحساب =======================
GROUPS = {
 "مستشفى عام": ["gd", "gs", "gt", "gp", "ga", "gadm"],
 "مستشفى ولادة وأطفال": ["md", "ms", "mt", "mp", "madm"],
 "مجمع إرادة والصحة النفسية": ["mh"],
 "مستشفى رعاية مديدة": ["lt", "la"],
}
mx = wb.create_sheet("مصفوفة الحساب"); mx.sheet_view.rightToLeft = True
h1(mx, "مصفوفة التجميع حسب الفئة ونوع المستشفى (ورقة حسابية وسيطة — لا تُعدَّل)", 9)
mx.cell(row=2, column=1, value="تجمع هذه الورقة نتائج أوراق المعايير حسب عمود «الفئة» لتغذية لوحة التحكم.").font = \
    Font(name=FONT, size=9, italic=True, color="595959")
for j, t in enumerate(TYPES):
    for off in (0, 4):
        cm = mx.cell(row=2, column=2 + j + off, value=t)
        cm.font = Font(name=FONT, size=8, color="BFBFBF")
        cm.alignment = Alignment(horizontal="center")
mx.cell(row=3, column=1, value="الفئة").font = HDR_FONT
mx.cell(row=3, column=1).fill = HDR_FILL
for j, t in enumerate(TYPES):
    for off, lab in ((0, "احتياج"), (4, "موجود")):
        c = mx.cell(row=3, column=2 + j + off, value=f"{t} — {lab}")
        c.font = HDR_FONT; c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center", wrap_text=True)
mx.row_dimensions[3].height = 40
for i, cat in enumerate(CATS):
    r = 4 + i
    mx.cell(row=r, column=1, value=cat).font = Font(name=FONT, size=10, bold=True)
    for j, t in enumerate(TYPES):
        for off, key in ((0, "col_need"), (4, "col_act")):
            parts = []
            for k in GROUPS[t]:
                m = meta[k]
                sn = m["sheet"]
                parts.append(
                    f"SUMIF('{sn}'!${m['col_cat']}${m['r0']}:${m['col_cat']}${m['rlast']},$A{r},"
                    f"'{sn}'!${m[key]}${m['r0']}:${m[key]}${m['rlast']})")
            c = mx.cell(row=r, column=2 + j + off, value="=" + "+".join(parts))
            c.number_format = "#,##0"; c.font = Font(name=FONT, size=10)
mx.column_dimensions["A"].width = 26
for j in range(2, 10): mx.column_dimensions[get_column_letter(j)].width = 15
mx.sheet_properties.tabColor = "A6A6A6"

# ======================= 5) ملخص لوحة التحكم =======================
dash.merge_cells("B13:H13")
c = dash.cell(row=13, column=2, value="② ملخص الاحتياج والفجوة (يُحدَّث تلقائياً)")
c.font = Font(name=FONT, size=12, bold=True, color=WHITE); c.fill = PatternFill("solid", fgColor=TEAL)
c.alignment = Alignment(horizontal="center"); dash.row_dimensions[13].height = 24

heads = ["الفئة", "الاحتياج حسب المعيار", "الموجود فعلياً", "الفجوة (+فائض / -عجز)",
         "نسبة التغطية", "المعدل لكل سرير", "الحالة"]
for j, t in enumerate(heads):
    c = dash.cell(row=14, column=2 + j, value=t)
    c.font = HDR_FONT; c.fill = PatternFill("solid", fgColor=NAVY); c.border = BOX
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
dash.row_dimensions[14].height = 36

CAT_LABEL = {
 "طبيب": "الأطباء", "تمريض": "التمريض", "قبالة": "القبالة",
 "أخصائي غير طبيب": "أخصائي غير طبيب (بدون تمريض وقبالة)",
 "فني": "الفنيون (بدون تمريض وقبالة)", "صيدلي": "الصيادلة",
 "مساعد صحي": "المساعد الصحي", "إداري": "الإداريون",
}
r0d = 15
for i, cat in enumerate(CATS):
    r = r0d + i
    dash.cell(row=r, column=2, value=CAT_LABEL[cat]).font = Font(name=FONT, size=10, bold=True)
    dash.cell(row=r, column=2).alignment = Alignment(horizontal="right")
    dash.cell(row=r, column=9, value=cat).font = Font(name=FONT, size=8, color="D9D9D9")  # مفتاح داخلي
    need = (f"=INDEX('مصفوفة الحساب'!$B$4:$E$11,MATCH($I{r},'مصفوفة الحساب'!$A$4:$A$11,0),"
            f"MATCH(HTYPE,'مصفوفة الحساب'!$B$2:$E$2,0))")
    act = (f"=INDEX('مصفوفة الحساب'!$F$4:$I$11,MATCH($I{r},'مصفوفة الحساب'!$A$4:$A$11,0),"
           f"MATCH(HTYPE,'مصفوفة الحساب'!$F$2:$I$2,0))")
    dash.cell(row=r, column=3, value=need).number_format = "#,##0"
    dash.cell(row=r, column=4, value=act).number_format = "#,##0"
    dash.cell(row=r, column=5, value=f"=D{r}-C{r}").number_format = "#,##0;[Red]-#,##0;-"
    dash.cell(row=r, column=6, value=f'=IF(C{r}=0,"",D{r}/C{r})').number_format = "0%"
    dash.cell(row=r, column=7, value=f'=IF(CAP=0,"",C{r}/CAP)').number_format = "0.00"
    dash.cell(row=r, column=8,
              value=f'=IF(C{r}=0,"لا ينطبق",IF(D{r}>=C{r},"مكتمل",IF(D{r}=0,"لم يُدخل","عجز")))')
    for j in range(2, 9):
        cc = dash.cell(row=r, column=j); cc.border = BOX
        if j > 2: cc.alignment = Alignment(horizontal="center")
        if cc.font.size is None or j > 2: cc.font = Font(name=FONT, size=10)
    dash.cell(row=r, column=2).font = Font(name=FONT, size=10, bold=True)

rT = r0d + len(CATS)
dash.cell(row=rT, column=2, value="إجمالي القوى العاملة")
for j, f in ((3, f"=SUM(C{r0d}:C{rT-1})"), (4, f"=SUM(D{r0d}:D{rT-1})"), (5, f"=D{rT}-C{rT}"),
             (6, f'=IF(C{rT}=0,"",D{rT}/C{rT})'), (7, f'=IF(CAP=0,"",C{rT}/CAP)')):
    dash.cell(row=rT, column=j, value=f)
dash.cell(row=rT, column=3).number_format = "#,##0"
dash.cell(row=rT, column=4).number_format = "#,##0"
dash.cell(row=rT, column=5).number_format = "#,##0;[Red]-#,##0;-"
dash.cell(row=rT, column=6).number_format = "0%"
dash.cell(row=rT, column=7).number_format = "0.00"
for j in range(2, 9):
    c = dash.cell(row=rT, column=j)
    c.font = Font(name=FONT, size=11, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=NAVY); c.border = BOX
    if j > 2: c.alignment = Alignment(horizontal="center")

dash.conditional_formatting.add(f"H{r0d}:H{rT-1}", FormulaRule(
    formula=[f'OR(H{r0d}="عجز",H{r0d}="لم يُدخل")'],
    fill=PatternFill("solid", fgColor="F8CBAD"), font=Font(name=FONT, color="833C0B")))
dash.conditional_formatting.add(f"H{r0d}:H{rT-1}", FormulaRule(
    formula=[f'H{r0d}="مكتمل"'],
    fill=PatternFill("solid", fgColor="C6E0B4"), font=Font(name=FONT, color="375623")))
dash.conditional_formatting.add(f"E{r0d}:E{rT}", CellIsRule(
    operator="lessThan", formula=["0"], font=Font(name=FONT, color="C00000", bold=True)))

# ملاحظة مصدر الأرقام
rn = rT + 2
dash.merge_cells(start_row=rn, start_column=2, end_row=rn, end_column=8)
c = dash.cell(row=rn, column=2,
    value="الاحتياج يُحتسب آلياً من أوراق المعايير حسب السعة السريرية المدخلة · «الموجود فعلياً» "
          "يُجمَّع من الخلايا الصفراء التي تعبّئها في أوراق المعايير · المعدل لكل سرير = الاحتياج ÷ السعة.")
c.font = Font(name=FONT, size=9, italic=True, color="595959")
c.alignment = Alignment(horizontal="right", wrap_text=True); dash.row_dimensions[rn].height = 28

dash.column_dimensions["A"].width = 3
dash.column_dimensions["B"].width = 34
for L, w in (("C", 18), ("D", 14), ("E", 18), ("F", 13), ("G", 15), ("H", 12)):
    dash.column_dimensions[L].width = w
dash.column_dimensions["I"].width = 3
dash.sheet_properties.tabColor = GOLD

# ======================= 6) حاسبة الأسرة والنسب =======================
cal = wb.create_sheet("حاسبة الأسرة والنسب"); cal.sheet_view.rightToLeft = True
h1(cal, "حاسبة توزيع الأسرة ونسب القوى العاملة", 6)
h2(cal, "مبنية على الإيضاحات 13-15 و 22-29 من الدليل — تعتمد السعة السريرية ونوع المستشفى من لوحة التحكم", 6)

def sect(ws, row, text, ncols=6):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=FONT, size=11, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=TEAL)
    c.alignment = Alignment(horizontal="center"); ws.row_dimensions[row].height = 22

def line(ws, row, label, formula, fmt="#,##0", note="", bold=False):
    c = ws.cell(row=row, column=1, value=label)
    c.font = Font(name=FONT, size=10, bold=bold)
    c.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
    c.border = BOX
    v = ws.cell(row=row, column=2, value=formula)
    v.number_format = fmt; v.fill = CALC_FILL; v.border = BOX
    v.font = Font(name=FONT, size=10, bold=True)
    v.alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
    n = ws.cell(row=row, column=3, value=note)
    n.font = Font(name=FONT, size=9, color="595959")
    n.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 22

GEN_T = '"مستشفى عام"'
MAT_T = '"مستشفى ولادة وأطفال"'

sect(cal, 4, "① توزيع أسرة العناية المركزة والطوارئ")
line(cal, 5, "السعة السريرية المدخلة", "=CAP", "#,##0", "من لوحة التحكم", True)
line(cal, 6, "إجمالي أسرة العناية المركزة (20% من السعة)", "=ROUND(CAP*20%,0)", "#,##0",
     "إيضاحان 13 و 14: نسبة أسرة العناية المركزة لجميع الفئات = 20% من السعة السريرية")
line(cal, 7, "أسرة عناية حديثي الولادة (NICU)",
     f'=ROUND(CAP*IF(HTYPE={MAT_T},13%,IF(CAP<=100,11%,10%)),0)', "#,##0",
     "عام: 10% (و 11% إذا كانت السعة 100 سرير فأقل) · ولادة وأطفال: 13%")
line(cal, 8, "أسرة عناية الكبار (ICU)",
     f'=ROUND(CAP*IF(HTYPE={MAT_T},0%,IF(CAP<=100,9%,7%)),0)', "#,##0",
     "عام: 7% (و 9% إذا كانت السعة 100 سرير فأقل) · لا تنطبق على مستشفيات الولادة والأطفال")
line(cal, 9, "أسرة عناية الأطفال (PICU)",
     f'=ROUND(CAP*IF(HTYPE={MAT_T},7%,IF(CAP<=100,0%,3%)),0)', "#,##0",
     "عام: 3% (ولا تُخصص عند 100 سرير فأقل) · ولادة وأطفال: 7%")
line(cal, 10, "أسرة قسم الطوارئ (حد أدنى 10%)", "=ROUNDUP(CAP*10%,0)", "#,##0",
     "إيضاح 15: لا تقل عن 10% من أسرة المستشفى ويمكن الزيادة بحسب عدد الزيارات")

sect(cal, 12, "② حاسبة التمريض (إيضاحات 11 و 28-30)")
line(cal, 13, "النسبة المعتمدة لكل سرير — الحد الأدنى", '=IF(MOBILEBEDS="نعم",1.68,2.84)', "0.00",
     "إيضاح 28: مع اعتماد الأسرة المتحركة 1.68-2.84 · إيضاح 29: بدونها 2.84-3.36")
line(cal, 14, "النسبة المعتمدة لكل سرير — الحد الأعلى", '=IF(MOBILEBEDS="نعم",2.84,3.36)', "0.00", "")
line(cal, 15, "إجمالي وظائف التمريض — الحد الأدنى", "=ROUNDUP(CAP*B13,0)", "#,##0", "")
line(cal, 16, "إجمالي وظائف التمريض — الحد الأعلى", "=ROUNDUP(CAP*B14,0)", "#,##0", "")
line(cal, 17, "إجمالي وظائف التمريض حسب جداول الدليل",
     "=INDEX('مصفوفة الحساب'!$B$4:$E$11,MATCH(\"تمريض\",'مصفوفة الحساب'!$A$4:$A$11,0),"
     "MATCH(HTYPE,'مصفوفة الحساب'!$B$2:$E$2,0))", "#,##0",
     "القيمة المعتمدة في هذا النموذج — مأخوذة من أوراق المعايير", True)
line(cal, 18, "فني تمريض (30%)", "=ROUND(B17*30%,0)", "#,##0",
     "إيضاح 11: نسبة مؤقتة حتى تقاعد الكادر الحالي ثم تُحدَّث تدريجياً وتُضاف إلى أخصائي")
line(cal, 19, "أخصائي تمريض (60% − نسبة الاستشاري)",
     "=ROUND(B17*(60%-IF(CAP>=300,0.5%,0)),0)", "#,##0",
     "إيضاح 30: عند أقل من 300 سرير تُضاف نسبة 0.5% إلى أخصائي تمريض")
line(cal, 20, "أخصائي أول تمريض (10%)", "=ROUND(B17*10%,0)", "#,##0", "")
line(cal, 21, "استشاري تمريض (0.5%)", "=IF(CAP>=300,ROUND(B17*0.5%,0),0)", "#,##0",
     "إيضاح 30: تبدأ من 300 سرير وما فوق")

sect(cal, 23, "③ حاسبة العلاج التنفسي (إيضاحات 22-25)")
line(cal, 24, "إجمالي القوى العاملة للعلاج التنفسي", "=ROUNDUP(CAP/5,0)", "#,##0",
     "إيضاح 25: تم أخذ المتوسط بواقع 1 أخصائي لكل 5 أسرة")
line(cal, 25, "فني علاج تنفسي (20%)", "=ROUND(B24*20%,0)", "#,##0", "إيضاح 22")
line(cal, 26, "أخصائي أول علاج تنفسي (20%)", "=ROUND(B24*20%,0)", "#,##0", "إيضاح 22")
line(cal, 27, "أخصائي علاج تنفسي (55%، أو 60% عند عدم وجود أسرة عناية)",
     "=ROUND(B24*IF(B6=0,60%,55%),0)", "#,##0", "إيضاح 23")
line(cal, 28, "استشاري علاج تنفسي (5%)", "=IF(B6=0,0,ROUND(B24*5%,0))", "#,##0",
     "إيضاح 23: فئة الاستشاري تكون في المستشفيات التي بها أسرة عناية مركزة فقط")
line(cal, 29, "إضافة عيادات العلاج التنفسي — عدد العيادات", 0, "#,##0",
     "إيضاح 24: أدخل عدد العيادات — يُؤمَّن 2 أخصائي لكل عيادة + أخصائي أول مشرف")
cal.cell(row=29, column=2).fill = IN_FILL; cal.cell(row=29, column=2).font = IN_FONT
line(cal, 30, "أخصائي علاج تنفسي إضافي للعيادات", "=B29*2", "#,##0", "إيضاح 24")
line(cal, 31, "أخصائي أول علاج تنفسي مشرف على العيادات", "=IF(B29>0,1,0)", "#,##0", "إيضاح 24")

sect(cal, 33, "④ مقارنة المعدلات لكل سرير مع المعدلات المنشورة في الدليل")
hdr = ["الفئة", "معدل مستشفاك (احتياج ÷ سعة)", "المعدل المحدّث المنشور", "التقييم", "", ""]
for j, t in enumerate(hdr[:4]):
    c = cal.cell(row=34, column=1 + j, value=t)
    c.font = HDR_FONT; c.fill = PatternFill("solid", fgColor=NAVY); c.border = BOX
    c.alignment = Alignment(horizontal="center", wrap_text=True)
cal.row_dimensions[34].height = 32
RATES = [
 ("طبيب", "الأطباء", 1.55, 1.55, "1.55 طبيب لكل سرير"),
 ("تمريض", "التمريض", 2.84, 3.36, "2.84 – 3.36 ممرضة لكل سرير"),
 ("قبالة", "القبالة", 0.20, 0.20, "0.20 قابلة لكل سرير"),
 ("صيدلي", "الصيادلة", 0.68, 0.68, "0.68 صيدلي لكل سرير"),
 ("إداري", "الإداريون", 0.68, 2.00, "0.68 – 2.00 لكل سرير حسب الفئة"),
]
for i, (key, lab, lo, hi, txt) in enumerate(RATES):
    r = 35 + i
    cal.cell(row=r, column=1, value=lab).font = Font(name=FONT, size=10, bold=True)
    cal.cell(row=r, column=1).border = BOX
    f = (f'=IF(CAP=0,"",INDEX(\'مصفوفة الحساب\'!$B$4:$E$11,'
         f'MATCH("{key}",\'مصفوفة الحساب\'!$A$4:$A$11,0),'
         f"MATCH(HTYPE,'مصفوفة الحساب'!$B$2:$E$2,0))/CAP)")
    c = cal.cell(row=r, column=2, value=f); c.number_format = "0.00"
    c.fill = CALC_FILL; c.border = BOX; c.alignment = Alignment(horizontal="center")
    c.font = Font(name=FONT, size=10, bold=True)
    c2 = cal.cell(row=r, column=3, value=txt); c2.border = BOX
    c2.font = Font(name=FONT, size=9); c2.alignment = Alignment(horizontal="center")
    c3 = cal.cell(row=r, column=4,
        value=f'=IF(B{r}="","",IF(B{r}<{lo},"أقل من المعدل",IF(B{r}>{hi},"أعلى من المعدل","ضمن المعدل")))')
    c3.border = BOX; c3.font = Font(name=FONT, size=10)
    c3.alignment = Alignment(horizontal="center")
cal.merge_cells(start_row=41, start_column=1, end_row=41, end_column=6)
c = cal.cell(row=41, column=1,
    value="تنبيه منهجي: المعدلات المنشورة في الدليل محسوبة على مجموع المستشفيات النموذجية السبعة "
          "(1700 سرير) وليست معدلاً لكل شريحة على حدة، لذلك قد يختلف معدل مستشفاك عنها دون أن يعني "
          "ذلك خللاً — القيمة المُلزِمة هي أعداد الجداول التفصيلية لا المعدل.")
c.font = Font(name=FONT, size=9, italic=True, color="C00000")
c.alignment = Alignment(horizontal="right", wrap_text=True); cal.row_dimensions[41].height = 40
cal.column_dimensions["A"].width = 46
cal.column_dimensions["B"].width = 20
cal.column_dimensions["C"].width = 30
for L in "DEF": cal.column_dimensions[L].width = 18
cal.sheet_properties.tabColor = TEAL

# ======================= 7) الطب الشرعي =======================
fs = wb.create_sheet("الطب الشرعي"); fs.sheet_view.rightToLeft = True
h1(fs, "معيار الطب الشرعي", 8)
h2(fs, "الملحق (1) — صفحة 31. محركات هذا المعيار مستقلة عن السعة السريرية: عدد عيون ثلاجات الموتى، عدد السكان، وعدد القضايا.", 8)

fs.cell(row=4, column=1, value="مدخلات المحركات (عبّئ الخلايا الصفراء)").font = Font(
    name=FONT, size=11, bold=True, color=WHITE)
fs.merge_cells("A4:H4"); fs.cell(row=4, column=1).fill = PatternFill("solid", fgColor=TEAL)
fs.cell(row=4, column=1).alignment = Alignment(horizontal="center")
drivers = [
 (5, "عدد عيون ثلاجات الموتى", 40),
 (6, "عدد السكان المخدومين (بالمليون نسمة)", 1.0),
 (7, "عدد قضايا الاشتباه بالكحول والمخدرات (سنوياً)", 14000),
 (8, "عدد قضايا الأحراز (سنوياً)", 13000),
 (9, "عدد قضايا السموم الإسعافية (سنوياً)", 2000),
]
for r, lab, val in drivers:
    c = fs.cell(row=r, column=1, value=lab)
    c.font = Font(name=FONT, size=10, bold=True); c.fill = PatternFill("solid", fgColor=LIGHT)
    c.border = BOX; c.alignment = Alignment(horizontal="right")
    v = fs.cell(row=r, column=2, value=val)
    v.fill = IN_FILL; v.font = IN_FONT; v.border = BOX
    v.number_format = "#,##0.0" if r == 6 else "#,##0"
    v.alignment = Alignment(horizontal="center")
fs.cell(row=5, column=3, value="◄ الشرائح: أقل من 20 · 20-50 · 51-150 · 151-250 · أكثر من 250").font = \
    Font(name=FONT, size=9, italic=True, color="7F7F7F")

def fhdr(row, cols):
    for j, t in enumerate(cols):
        c = fs.cell(row=row, column=1 + j, value=t)
        c.font = HDR_FONT; c.fill = PatternFill("solid", fgColor=NAVY); c.border = BOX
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    fs.row_dimensions[row].height = 30

# أ) إدارة الوفيات
fs.cell(row=11, column=1, value="أ) إدارة الوفيات — المحرك: عدد عيون ثلاجات الموتى").font = \
    Font(name=FONT, size=11, bold=True, color=NAVY)
fhdr(12, ["المسمى الوظيفي", "أقل من 20", "20 - 50", "51 - 150", "151 - 250", "أكثر من 250",
          "الاحتياج المحسوب", "الموجود فعلياً"])
for i, (name, vals, tot) in enumerate(ml.FORENSIC_MORTUARY):
    r = 13 + i
    fs.cell(row=r, column=1, value=name).font = Font(name=FONT, size=10)
    fs.cell(row=r, column=1).border = BOX
    for j, v in enumerate(vals):
        c = fs.cell(row=r, column=2 + j, value=v)
        c.fill = TIER_FILL; c.border = BOX; c.number_format = "#,##0"
        c.alignment = Alignment(horizontal="center"); c.font = Font(name=FONT, size=10)
    c = fs.cell(row=r, column=7,
        value=f"=INDEX(B{r}:F{r},IF($B$5<20,1,IF($B$5<=50,2,IF($B$5<=150,3,IF($B$5<=250,4,5)))))")
    c.fill = CALC_FILL; c.border = BOX; c.number_format = "#,##0"
    c.font = Font(name=FONT, size=10, bold=True); c.alignment = Alignment(horizontal="center")
    a = fs.cell(row=r, column=8); a.fill = IN_FILL; a.font = IN_FONT; a.border = BOX
    a.number_format = "#,##0"; a.alignment = Alignment(horizontal="center")

# ب) إدارة الطب الشرعي — لكل مليون نسمة
fs.cell(row=16, column=1, value="ب) إدارة الطب الشرعي — المحرك: عدد السكان (لكل مليون نسمة)").font = \
    Font(name=FONT, size=11, bold=True, color=NAVY)
fhdr(17, ["المسمى الوظيفي", "المعدل لكل مليون نسمة", "الاحتياج المحسوب", "الموجود فعلياً"])
r = 18
for name, rate in ml.FORENSIC_POP:
    fs.cell(row=r, column=1, value=name).font = Font(name=FONT, size=10)
    fs.cell(row=r, column=1).border = BOX
    c = fs.cell(row=r, column=2, value=rate); c.fill = TIER_FILL; c.border = BOX
    c.alignment = Alignment(horizontal="center"); c.number_format = "#,##0"
    c2 = fs.cell(row=r, column=3, value=f"=ROUNDUP(B{r}*$B$6,0)")
    c2.fill = CALC_FILL; c2.border = BOX; c2.number_format = "#,##0"
    c2.font = Font(name=FONT, size=10, bold=True); c2.alignment = Alignment(horizontal="center")
    a = fs.cell(row=r, column=4); a.fill = IN_FILL; a.font = IN_FONT; a.border = BOX
    a.number_format = "#,##0"; a.alignment = Alignment(horizontal="center")
    r += 1
fs.cell(row=r, column=1, value="فني تمريض (2 لكل استشاري طبيب شرعي)").font = Font(name=FONT, size=10)
fs.cell(row=r, column=1).border = BOX
fs.cell(row=r, column=2, value=2).fill = TIER_FILL
fs.cell(row=r, column=2).alignment = Alignment(horizontal="center")
fs.cell(row=r, column=2).border = BOX
c2 = fs.cell(row=r, column=3, value=f"=C20*2"); c2.fill = CALC_FILL; c2.border = BOX
c2.number_format = "#,##0"; c2.font = Font(name=FONT, size=10, bold=True)
c2.alignment = Alignment(horizontal="center")
a = fs.cell(row=r, column=4); a.fill = IN_FILL; a.font = IN_FONT; a.border = BOX
fs.cell(row=r + 1, column=1,
    value="ملاحظة تدقيق: مجموع هذا القسم المنشور في الدليل = 11، بينما مجموع المعدلات (2+3+4) = 9؛ "
          "الفارق يعادل بند فني تمريض المحتسب مرة أخرى في صفه المستقل.").font = \
    Font(name=FONT, size=8, italic=True, color="C00000")
fs.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=8)
fs.cell(row=r + 1, column=1).alignment = Alignment(horizontal="right", wrap_text=True)

# ج) إدارة السموم
blocks = [
 (24, "ج) إدارة السموم — حالات الاشتباه بالكحول والمخدرات (لكل 14 ألف قضية)", ml.FORENSIC_TOX_14K, 14000, "$B$7"),
 (33, "د) إدارة السموم — قضايا الأحراز (لكل 13 ألف قضية)", ml.FORENSIC_TOX_13K, 13000, "$B$8"),
 (42, "هـ) إدارة السموم — السموم الإسعافية (لكل ألفي قضية)", ml.FORENSIC_TOX_2K, 2000, "$B$9"),
]
for start, title, rows, per, drv in blocks:
    fs.cell(row=start, column=1, value=title).font = Font(name=FONT, size=11, bold=True, color=NAVY)
    fhdr(start + 1, ["المسمى الوظيفي", f"المعدل لكل {per:,} قضية", "الاحتياج المحسوب", "الموجود فعلياً"])
    for i, (name, rate) in enumerate(rows):
        r = start + 2 + i
        fs.cell(row=r, column=1, value=name).font = Font(name=FONT, size=10)
        fs.cell(row=r, column=1).border = BOX
        c = fs.cell(row=r, column=2, value=rate); c.fill = TIER_FILL; c.border = BOX
        c.alignment = Alignment(horizontal="center")
        c2 = fs.cell(row=r, column=3, value=f"=ROUNDUP(B{r}*{drv}/{per},0)")
        c2.fill = CALC_FILL; c2.border = BOX; c2.number_format = "#,##0"
        c2.font = Font(name=FONT, size=10, bold=True); c2.alignment = Alignment(horizontal="center")
        a = fs.cell(row=r, column=4); a.fill = IN_FILL; a.font = IN_FONT; a.border = BOX
        a.number_format = "#,##0"; a.alignment = Alignment(horizontal="center")
fs.column_dimensions["A"].width = 44
for L in "BCDEFGH": fs.column_dimensions[L].width = 15
fs.sheet_properties.tabColor = TEAL

# ======================= 8) الإيضاحات =======================
NOTES = [
 "هذا المعيار هو الحد الأدنى للقوى العاملة.",
 "في المستشفيات التي لا يوجد بها استشاريون أو أخصائيون في التخصصات الدقيقة يمكن الاستفادة من المستشفى الافتراضي، والاستشارات الافتراضية، والاستشارات على 937.",
 "يفضل العمل في الأقسام بطريقة (General Medical Team – GMT) وهو فريق استشاريين للباطنة يعمل في وحدة الباطنة ويتكون الفريق من 2 استشاري و2 أخصائي و4 مقيم، مع التنويه بأن عدد العيادات لكل استشاري لا تقل عن 4 عيادات أسبوعياً بالإضافة للعيادة الافتراضية.",
 "في حال وجود طبيب لتخصص دقيق غير مذكور يتم توظيفهم على التخصص العام.",
 "يتوجب وجود فريقين على الأقل بالمستشفيات التي أقل من أو يساوي 200 سرير، وأربع فرق في أكثر من 200 سرير.",
 "في حال النقص يتم الرفع بطلب السعة السريرية أو طلب توظيف أو استخدام المستشفى الافتراضي أو التعاقد الجزئي.",
 "وظائف الاستشاريين والأخصائيين والمقيمين لجميع التخصصات يتم تحديدها فقط للقسم المخصص لها. وظائف الجودة ومكافحة العدوى والأقسام الأخرى يتم احتساب قوى عاملة خاصة لها.",
 "يطبق معيار العنايات المركزة إذا كانت المستشفى تقدم الخدمة ومعتمدة.",
 "الطبيب المقيم فئة عامة دون تخصص وتم تخصيصه بمسمى معين لإلزامه بعمل معين.",
 "معيار القابلات خاص بالمستشفيات العامة التي تقدم خدمات النساء والولادة ومستشفيات النساء والولادة التخصصية، ويتم توزيع القابلات على أقسام (ما قبل الولادة – ما بعد الولادة – الطوارئ – غرف الولادة).",
 "يتم شغل وظائف فني بنسبة 30% إلى حين تقاعد الكادر الحالي أو تحسين وضع الموظفات، ثم يتم تحديث النسبة تدريجياً وإضافتها إلى أخصائي.",
 "تقتصر خدمات طب الأسنان في المستشفيات العامة على وحدة جراحة الفم والوجه والفكين لخدمة المرضى المنومين في المستشفى.",
 "في المستشفيات العامة تم حساب القوى العاملة للعنايات المركزة بناءً على نسبة عدد أسرة العناية المركزة لجميع الفئات وهي 20% من السعة السريرية وتقسم كالتالي: 10% أسرة أطفال حديثي الولادة، 7% أسرة عناية للكبار، 3% أسرة عناية للأطفال. وفي المستشفيات ذات السعة 100 سرير وأقل تقسم كالتالي: 11% حديثي الولادة، 9% عناية للكبار.",
 "في مستشفيات الولادة والأطفال تم حساب القوى العاملة للعنايات المركزة بناءً على نسبة 20% من السعة السريرية وتقسم كالتالي: 7% أسرة عناية أطفال، 13% أسرة أطفال حديثي ولادة.",
 "عدد أسرة قسم الطوارئ لا تقل عن 10% من أسرّة المستشفى ويمكن الزيادة بحسب عدد زيارات المرضى.",
 "عدد الأطباء في قسم الطوارئ (استشاريين وأخصائيين وأطباء مقيمين) لابد أن يكون كافياً لتغطية العمل على مدار الساعة بنظام الورديات وبحسب عدد زيارات المرضى ومستوى خطورتها وتعقيدها.",
 "لابد من الأخذ في الاعتبار العدد القادر على تغطية العمل في قسم الطوارئ في الأوقات والأيام عالية الضغط وكذلك لتغطية العجز (مرض أو ظروف قاهرة وغيرها).",
 "في كل وردية لابد من وجود استشاري أو أخصائي حاصل على شهادة الاختصاص في طب الطوارئ في أقسام طوارئ الكبار، واستشاري أو أخصائي حاصل على زمالة طوارئ الأطفال في أقسام طوارئ الأطفال.",
 "يتم اعتماد معيار تخصصات الجراحة العامة في حال كان المستشفى مرجعياً.",
 "يتم اعتماد معيار العظام في حال كان المستشفى مرجعياً.",
 "الفئات الطبية المساعدة تشمل أخصائي غير طبيب وفني بدون تمريض ومساعد صحي.",
 "يتم توزيع أعداد القوى العاملة للعلاج التنفسي من حيث التصنيف على النحو التالي: 20% فني، 20% أخصائي أول، 55% أخصائي، 5% استشاري.",
 "فئة استشاري علاج تنفسي تكون في المستشفيات التي يوجد بها أسرة عناية مركزة، وفي حال عدم وجود أسرة عناية مركزة تضاف النسبة إلى فئة الأخصائيين حيث تكون النسبة 60% من إجمالي أعداد القوى العاملة للأخصائيين.",
 "في حال وجود عيادات خاصة بالعلاج التنفسي يتم تأمين عدد 2 أخصائي علاج تنفسي لكل عيادة، كذلك أخصائي أول علاج تنفسي كمشرف على العيادات.",
 "تم أخذ المتوسط لعدد الأسرة لكل أخصائي علاج تنفسي بـ 1:5.",
 "الأسرة المتحركة هي (العيادات – الطوارئ – أقسام الولادة – الأقسام الفنية الإدارية).",
 "الأقسام الفنية الإدارية هي (التدريب والتعليم – مكافحة العدوى – الجودة التمريضية – التعقيم – إلخ).",
 "المعيار للتمريض في حال اعتماد الأسرة المتحركة يبقى كما هو 1.68 – 2.84.",
 "المعيار للتمريض في حال عدم اعتماد الأسرة المتحركة يصبح 2.84 – 3.36.",
 "أخصائي استشاري تمريض 0.5% تبدأ من 300 سرير وما فوق، وفي الأقل تضاف 0.5% لمسمى أخصائي تمريض.",
 "عند تخريج أو فتح معاهد أو كليات للفنيين يتم الزيادة تدريجياً لحين الوصول للمستهدف.",
 "تم تعديل مسمى كاتب وكاتب جناح وموظف استقبال ومشغل حاسب آلي إلى مسمى مساعد إداري توافقاً مع الدليل السعودي الموحد للمهن.",
 "لا يوجد اختلاف من حيث العدد في قسم الأشعة، وإنما التحديث في إدخال بعض المسميات التقنية لوجود مخرجات من الجامعات الداخلية والخارجية وبرامج الزمالات لتلك الدرجات، وعليه تم إعادة توزيع الأعداد فقط.",
 "يتم اعتماد معيار التأهيل الطبي في المستشفيات التي يتوفر بها مركز تأهيل معتمد.",
 "في حال تقديم المستشفى لخدمات الأطفال يضاف طبيب مسالك بولية أطفال عدد (1) لسعة 400/500 سرير.",
 "في المستشفيات التخصصية 500 سرير وما فوق تكون القوى العاملة للتخصصات الدقيقة التابعة للمسالك البولية على النحو التالي: 2 استشاري لكل تخصص، 1 أخصائي لكل 2 استشاري، 4 مقيم لكل استشاري.",
 "تطبق المسميات من 21 إلى 30 في حال توفّر خدمات الأورام في المنشأة.",
 "يطبق معيار استشاري زراعة القوقعة في حال اعتماد مركز متخصص لزراعة القوقعة.",
 "يطبق معيار استشاري عقم وأطفال أنابيب في حال اعتماد مركز للعقم في المنشأة.",
 "تم إضافة فني رعاية مرضى (مساعدة ممرضة سابقاً) بناءً على معايير وتعميم الهيئة السعودية للتخصصات الصحية.",
 "يطبق معيار جراحة السمنة في حال كان المركز معتمداً.",
 "يشترط لتطبيق المعيار توفّر نطاق الخدمة في المنشأة.",
]
nsh = wb.create_sheet("الإيضاحات"); nsh.sheet_view.rightToLeft = True
h1(nsh, "الإيضاحات الواردة في الدليل (42 إيضاحاً)", 3)
h2(nsh, "الملحق (1) — صفحتا 32 و 33. هذه الإيضاحات ملزمة عند تطبيق أي رقم في هذا النموذج.", 3)
nsh.cell(row=4, column=1, value="م").font = HDR_FONT
nsh.cell(row=4, column=1).fill = HDR_FILL
nsh.cell(row=4, column=2, value="الإيضاح").font = HDR_FONT
nsh.cell(row=4, column=2).fill = HDR_FILL
nsh.merge_cells("B4:C4")
for i, t in enumerate(NOTES):
    r = 5 + i
    c = nsh.cell(row=r, column=1, value=i + 1)
    c.font = Font(name=FONT, size=10, bold=True, color=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center"); c.border = BOX
    nsh.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    c = nsh.cell(row=r, column=2, value=t)
    c.font = Font(name=FONT, size=10)
    c.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True); c.border = BOX
    nsh.row_dimensions[r].height = max(28, 15 * (len(t) // 95 + 1))
nsh.column_dimensions["A"].width = 6
nsh.column_dimensions["B"].width = 90
nsh.column_dimensions["C"].width = 30
nsh.sheet_properties.tabColor = GOLD

# ======================= 9) نطاق الخدمات =======================
SCOPE = {
 "50 سرير": ["طب الباطنة","رعاية منزلية","كلى","غدد صماء وسكري","أمراض الصدرية","طب نفسي",
   "أمراض جلدية وتناسلية","جراحة عامة","مسالك بولية","عظام","عيون","أنف وأذن وحنجرة","أطفال",
   "نساء وولادة","تخدير","عناية مركزة كبار","عناية مركزة حديثي ولادة","طوارئ","أشعة","أسنان",
   "مكافحة عدوى","جودة","باثولوجيا إكلينيكية"],
 "100 سرير": ["طب الباطنة","رعاية منزلية","قلب","جهاز هضمي وكبد","كلى","غدد صماء وسكري","أورام كبار",
   "أمراض الصدرية","أمراض معدية","طب نفسي","طب نفسي أطفال ومراهقين","أمراض جلدية وتناسلية",
   "جراحة عامة","مسالك بولية","عظام","عيون","أنف وأذن وحنجرة","سمعيات","جراحة مخ وأعصاب","أطفال",
   "نساء وولادة","تخدير","عناية مركزة كبار","عناية مركزة حديثي ولادة","طوارئ","أشعة",
   "جراحة وجه وفكين","أسنان","مكافحة عدوى","جودة","باثولوجيا إكلينيكية","هيماتولوجي","كيمياء حيوية",
   "بنك دم","جراثيم"],
 "150 سرير": ["طب الباطنة","رعاية منزلية","قلب","جهاز هضمي وكبد","أعصاب","كلى","غدد صماء وسكري",
   "أمراض دم وأورام","أورام كبار","علاج إشعاعي أورام","علاج تلطيفي أورام","أورام أطفال",
   "أمراض الصدرية","أمراض معدية","طب نفسي","طب نفسي أطفال ومراهقين","أمراض جلدية وتناسلية",
   "جراحة عامة","مسالك بولية","عظام","عيون","أنف وأذن وحنجرة","سمعيات","أمراض تخاطب",
   "جراحة مخ وأعصاب","جراحة أوعية دموية","جراحة تجميل وحروق","أمراض السمنة","جراحة أطفال","أطفال",
   "قلب أطفال","نساء وولادة","تخدير","عناية مركزة كبار","عناية مركزة أطفال","عناية مركزة حديثي ولادة",
   "طوارئ","أشعة","جراحة وجه وفكين","أسنان","مكافحة عدوى","جودة","باثولوجيا إكلينيكية","هيماتولوجي",
   "كيمياء حيوية","بنك دم","جراثيم"],
 "200 سرير": ["طب الباطنة","رعاية منزلية","قلب","جهاز هضمي وكبد","أعصاب","كلى","كلى أطفال",
   "غدد صماء وسكري","أمراض دم وأورام","أورام كبار","علاج إشعاعي أورام","علاج تلطيفي أورام",
   "أورام أطفال","أمراض وراثية واستقلابية","أمراض الصدرية","أمراض معدية","روماتيزم","طب نفسي",
   "طب نفسي أطفال ومراهقين","علم المناعة والحساسية","أمراض جلدية وتناسلية","طب طبيعي وتأهيل",
   "جراحة عامة","جراحة سمنة","مسالك بولية","عظام","عيون","أنف وأذن وحنجرة","سمعيات","أمراض تخاطب",
   "جراحة مخ وأعصاب","جراحة صدر","جراحة أوعية دموية","جراحة تجميل وحروق","أمراض السمنة",
   "جراحة أطفال","أطفال","قلب أطفال","نمو وسلوك أطفال","نساء وولادة","عقم وأطفال أنابيب","تخدير",
   "عناية مركزة كبار","عناية مركزة أطفال","عناية مركزة حديثي ولادة","طوارئ","أشعة","جراحة وجه وفكين",
   "أسنان","مكافحة عدوى","جودة","باثولوجيا إكلينيكية","هيماتولوجي","كيمياء حيوية","بنك دم","جراثيم",
   "تشريح نسيجي"],
 "300 سرير": ["طب الباطنة","رعاية منزلية","قلب","جهاز هضمي وكبد","أعصاب","كلى","كلى أطفال",
   "غدد صماء وسكري","أمراض دم وأورام","أورام كبار","علاج إشعاعي أورام","علاج تلطيفي أورام",
   "أورام أطفال","أمراض وراثية واستقلابية","أمراض الصدرية","أمراض معدية","روماتيزم","طب نفسي",
   "طب نفسي أطفال ومراهقين","علم المناعة والحساسية","أمراض جلدية وتناسلية","طب طبيعي وتأهيل",
   "جراحة عامة","جراحة غدد صماء وثدي","جراحة كبد وقنوات مرارية","جراحة قولون ومستقيم","جراحة سمنة",
   "جراحة جهاز هضمي","مسالك بولية","عظام","عيون","أنف وأذن وحنجرة","سمعيات","زراعة القوقعة",
   "أمراض تخاطب","جراحة مخ وأعصاب","جراحة عمود فقري","جراحة صدر","جراحة أوعية دموية",
   "جراحة تجميل وحروق","أمراض السمنة","جراحة أطفال","أطفال","قلب أطفال","أعصاب أطفال",
   "نمو وسلوك أطفال","نساء وولادة","عقم وأطفال أنابيب","تخدير","عناية مركزة كبار","عناية مركزة أطفال",
   "عناية مركزة حديثي ولادة","طوارئ","أشعة","طب نووي","طب كبار السن (شيخوخة)","جراحة وجه وفكين",
   "أسنان","مكافحة عدوى","جودة","باثولوجيا إكلينيكية","هيماتولوجي","كيمياء حيوية","بنك دم","جراثيم",
   "فيروسات","طفيليات","تشريح نسيجي"],
}
SCOPE["400 / 500 سرير"] = SCOPE["300 سرير"] + ["مختبر مناعة"]
sc = wb.create_sheet("نطاق الخدمات"); sc.sheet_view.rightToLeft = True
ncol = len(SCOPE)
h1(sc, "نطاق الخدمات المعتمد لكل سعة سريرية (المستشفيات العامة)", ncol)
h2(sc, "الملحق (1) — صفحات 41-44. إيضاح 42: يشترط لتطبيق المعيار توفّر نطاق الخدمة في المنشأة.", ncol)
for j, (k, items) in enumerate(SCOPE.items(), start=1):
    c = sc.cell(row=4, column=j, value=k)
    c.font = HDR_FONT; c.fill = PatternFill("solid", fgColor=NAVY); c.border = BOX
    c.alignment = Alignment(horizontal="center")
    for i, it in enumerate(items):
        cc = sc.cell(row=5 + i, column=j, value=f"{i+1}. {it}")
        cc.font = Font(name=FONT, size=9); cc.border = BOX
        cc.alignment = Alignment(horizontal="right")
    cnt = sc.cell(row=5 + len(SCOPE["400 / 500 سرير"]) + 1, column=j, value=len(items))
    cnt.font = Font(name=FONT, size=10, bold=True, color=WHITE)
    cnt.fill = PatternFill("solid", fgColor=TEAL); cnt.number_format = '#,##0" خدمة"'
    cnt.alignment = Alignment(horizontal="center")
for j in range(1, ncol + 1):
    sc.column_dimensions[get_column_letter(j)].width = 27
sc.freeze_panes = "A5"
sc.sheet_properties.tabColor = GOLD

# ======================= 10) التحقق من المطابقة =======================
vs = wb.create_sheet("التحقق من المطابقة"); vs.sheet_view.rightToLeft = True
h1(vs, "تدقيق مطابقة النموذج لأرقام الدليل", 6)
h2(vs, "لكل ورقة معيار: مجموع الصفوف المُدخلة في النموذج مقابل المجموع المنشور في الدليل عند كل شريحة سعة.", 6)
for j, t in enumerate(["ورقة المعيار", "عدد المسميات", "المجموع الكلي (محسوب)",
                       "المجموع الكلي (منشور)", "الفرق", "النتيجة"]):
    c = vs.cell(row=4, column=1 + j, value=t)
    c.font = HDR_FONT; c.fill = PatternFill("solid", fgColor=NAVY); c.border = BOX
    c.alignment = Alignment(horizontal="center", wrap_text=True)
vs.row_dimensions[4].height = 32
PUBLISHED_GRAND = {
 "عام - الأطباء": sum(gd.PUBLISHED_DOCTORS_TOTALS),
 "عام - أخصائي غير طبيب": sum(gr.PUB_SPEC_TOTAL),
 "عام - الفنيون": sum(gr.PUB_TECH_TOTAL),
 "عام - الصيادلة": sum(gr.PUB_PHARM_TOTAL),
 "عام - المساعد الصحي": sum(gr.PUB_AUX_TOTAL),
 "عام - الإداريون": sum(gr.PUB_ADMIN_TOTAL),
 "ولادة - الأطباء": sum(mt.PUB_DOCTORS_M),
 "ولادة - الفنيون": sum(mt.PUB_TECH_M_TOTAL),
 "ولادة - الصيادلة": sum(mt.PUB_PHARM_M),
 "ولادة - الإداريون": sum(mt.PUB_ADMIN_M),
 "الصحة النفسية وإرادة": sum(ml.PUB_MH_TOTAL),
 "الرعاية المديدة - الفنية": sum(ml.PUB_LTC_TECH),
 "الرعاية المديدة - الإداريون": sum(ml.PUB_LTC_ADMIN),
}
ORDER = ["gd","gs","gt","gp","ga","gadm","md","ms","mt","mp","madm","mh","lt","la"]
r = 5
for k in ORDER:
    m = meta[k]; sn = m["sheet"]; n = len(m["tiers"])
    t0 = get_column_letter(4); t1 = get_column_letter(3 + n)
    vs.cell(row=r, column=1, value=sn).font = Font(name=FONT, size=10, bold=True)
    vs.cell(row=r, column=2, value=m["rlast"] - m["r0"] + 1).number_format = "#,##0"
    vs.cell(row=r, column=3, value=f"=SUM('{sn}'!${t0}${m['r0']}:${t1}${m['rlast']})")
    pub = PUBLISHED_GRAND.get(sn, "")
    vs.cell(row=r, column=4, value=pub)
    vs.cell(row=r, column=5, value=f'=IF(D{r}="","",C{r}-D{r})')
    vs.cell(row=r, column=6,
            value=f'=IF(D{r}="","لا يوجد مجموع منشور مستقل",IF(E{r}=0,"مطابق تماماً","فرق يستوجب المراجعة"))')
    for j in range(1, 7):
        c = vs.cell(row=r, column=j); c.border = BOX
        if j > 1:
            c.alignment = Alignment(horizontal="center")
            c.font = Font(name=FONT, size=10)
        if j in (3, 4, 5): c.number_format = "#,##0"
    r += 1
vs.conditional_formatting.add(f"F5:F{r-1}", FormulaRule(
    formula=['F5="مطابق تماماً"'], fill=PatternFill("solid", fgColor="C6E0B4"),
    font=Font(name=FONT, color="375623")))
vs.conditional_formatting.add(f"F5:F{r-1}", FormulaRule(
    formula=['F5="فرق يستوجب المراجعة"'], fill=PatternFill("solid", fgColor="F8CBAD"),
    font=Font(name=FONT, color="833C0B")))
r += 1
findings = [
 ("ملاحظات التدقيق على الدليل المصدر", ""),
 ("ورقة «ولادة - الفنيون»",
  "جدول الفنيين في الملحق (2) يتضمن صف «فني ترميز طبي» (2/3/3/3/4/4) دون رقم تسلسلي، وهو غير مُدرج "
  "في المجموع المنشور (186/270/355/510/680/844). النموذج يحتسبه ضمن المجموع، ولذلك يظهر فرق ثابت قدره 19 وظيفة."),
 ("ورقة «ولادة - الفنيون» — فني قبالة",
  "القيمة عند 300 سرير (12) تخالف تدرّج بقية الشرائح (2/3/4/…/6/10)؛ أُبقيت كما وردت في الدليل حرفياً."),
 ("ورقة «الطب الشرعي»",
  "المجموع المنشور لقسم إدارة الطب الشرعي = 11 بينما مجموع المعدلات (2+3+4) = 9؛ الفارق يعادل بند فني تمريض."),
 ("جدول الملخص التنفيذي (صفحة 21)",
  "أرقام ملخص الأطباء في صفحة 21 (804/971/820 = 2595) تختلف قليلاً عن الجدول التفصيلي في صفحات 22-25 "
  "(807/980/820 = 2607). النموذج يعتمد الجدول التفصيلي لأنه المصدر الملزم."),
 ("منهجية المعدلات",
  "المعدلات المنشورة «لكل سرير» محسوبة على مجموع المستشفيات النموذجية السبعة (1700 سرير) وليست معدلاً "
  "لكل شريحة، وهو ما يفسّر اختلافها عن معدل أي مستشفى منفرد."),
]
for i, (a, b) in enumerate(findings):
    rr = r + i
    if b == "":
        vs.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=6)
        c = vs.cell(row=rr, column=1, value=a)
        c.font = Font(name=FONT, size=11, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=TEAL); c.alignment = Alignment(horizontal="center")
    else:
        vs.cell(row=rr, column=1, value=a).font = Font(name=FONT, size=9, bold=True, color=NAVY)
        vs.merge_cells(start_row=rr, start_column=2, end_row=rr, end_column=6)
        c = vs.cell(row=rr, column=2, value=b)
        c.font = Font(name=FONT, size=9)
        c.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
        vs.row_dimensions[rr].height = 40
vs.column_dimensions["A"].width = 30
for L, w in (("B", 24), ("C", 20), ("D", 20), ("E", 14), ("F", 26)):
    vs.column_dimensions[L].width = w
vs.sheet_properties.tabColor = "A6A6A6"

# ======================= ترتيب الأوراق وحفظ الملف =======================
order = ["دليل الاستخدام", "لوحة التحكم",
         "عام - الأطباء", "عام - أخصائي غير طبيب", "عام - الفنيون", "عام - الصيادلة",
         "عام - المساعد الصحي", "عام - الإداريون",
         "ولادة - الأطباء", "ولادة - أخصائي غير طبيب", "ولادة - الفنيون", "ولادة - الصيادلة",
         "ولادة - الإداريون",
         "الصحة النفسية وإرادة", "الرعاية المديدة - الفنية", "الرعاية المديدة - الإداريون",
         "الطب الشرعي", "حاسبة الأسرة والنسب", "نطاق الخدمات", "الإيضاحات",
         "التحقق من المطابقة", "مصفوفة الحساب"]
wb._sheets = [wb[s] for s in order]
wb.active = 1
# لا تُخزَّن قيم محسوبة داخل الملف (openpyxl لا يكتبها)، لذا يُطلب من Excel
# إعادة حساب الملف بالكامل عند كل فتح.
wb.calculation.fullCalcOnLoad = True
wb.save(OUT)
print("saved:", OUT)
print("sheets:", len(wb.sheetnames))
